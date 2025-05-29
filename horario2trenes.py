# ESTRUCTURA EXCEL : CADA TREN SE PONE EN UNA HOJA DIFERENTE, EN LA PRIMERA FILA SE PONEN LOS LIMITES DE DONDE SE DEBE LEER
#							 minX1 | maxX1 | minY1 | maxY1 | minX2 | maxX2 | minY2 | maxY2

import json
import copy
import openpyxl
from tqdm import tqdm
from openpyxl.styles import Font
from json.decoder import JSONDecodeError

# ============ EN DESARROLLO ============

# ============ MÁS TARDE ============

# --------- CLASES ---------

class Ciudad:
	def __init__(self, id, nombre, coords):
		self.id = id
		self.nombre = nombre
		self.coords = coords

class Punto:
	def __init__(self, coords, parada, id_ciudad): # Cuando es usado para ruta, parada siempre es False y ciudad = None si no hay ciudad
		self.coords = coords
		self.parada = parada
		self.id_ciudad = id_ciudad

class Ruta:
	def __init__(self, id, puntos):
		self.id = id
		self.puntos = puntos

class Trayecto:
	def __init__(self, puntos, salidas, llegadas):
		self.puntos = puntos
		
		self.llegadas = llegadas
		self.salidas = salidas

class Linea:
	def __init__(self, id, nombre, color, trayectos, rutas):
		self.id = id
		self.nombre = nombre
		self.color = color
		
		self.trayectos = trayectos
		self.rutas = rutas

# --------- VARIABLES ---------

horariosF = "Horarios2025.xlsx" # Nombre del archivo en el que se leen los horarios

ciudadesF = "Ciudades.json" # Nombre des archivo en el que se leen las ciudades para verificar
lineasF = "LineasIC.json" # Nombre del archivo en el que se leen y guardan los trenes

ciudades = [] # ciudades como en el editor

lineas = [] # lineas como en el editor
lineas_escribir = [] # lineas para ser escritas

# --------- FUNCIONES ---------

def getCiudadfromId(id): 	# Devuelve la ciudad desde una id
	for c in ciudades:
		if (c.id == id):
			return c
	return None

def getPosParadaRuta(idP, ruta):	# Devuelve la posicion de la parada con id "idP" en la ruta "ruta" (trayecto[1]) 
	for r in range(len(ruta)):
		if (idP == ruta[r].id_ciudad): return r
	
	return None

def idC2ciudades(ids): 		# Recibe un array con ids de ciudades y devuelve un array con los nombres de las ciudades
	temp = []
	for id in ids:
		c = getCiudadfromId(id)
		if (c == None):
			print("Error inesperado, el id de una parada de tren no existe")
			return None
		
		temp.append(c.nombre)
	return temp

def int2hora(horaInt):
	if(horaInt == None): return None
	temp = str(horaInt)
	return ('0' * (4 - len(temp))) + temp

# --- Cargar datos ---

def cargar_ciudades():
	global ciudadesF, ciudades
	
	try:
		f = open(ciudadesF, 'r')
		data = json.loads(f.read())
		
		for i in data: # Cargar ciudades
			ciudades.append( Ciudad(i['id'], i['nombre'], i['coords']) )
		
	except JSONDecodeError:
		pass

def cargar_lineas():
	global lineas, lineasF, linea_actual
	
	with open(lineasF, "r") as f:
		data = json.load(f)
		
		for linea_dict in data:
			id_linea = linea_dict["id"]
			nombre = linea_dict["nombre"]
			color = linea_dict["color"]

			# Parse trayectos
			trayectos = []
			for trayecto_dict in linea_dict["trayectos"]:
				puntos = [
					Punto(p["coords"], p["parada"], p.get("id_ciudad"))
					for p in trayecto_dict["puntos"]
				]
				salidas = trayecto_dict["salidas"]
				llegadas = trayecto_dict["llegadas"]
				trayectos.append(Trayecto(puntos, salidas, llegadas))

			# Parse rutas
			rutas = []
			for ruta_dict in linea_dict["rutas"]:
				puntos = [
					Punto(p["coords"], p["parada"], p.get("id_ciudad"))
					for p in ruta_dict["puntos"]
				]
				rutas.append(Ruta(ruta_dict["id"], puntos))
			
			lineas.append(Linea(id_linea, nombre, color, trayectos, rutas))

# --- Verificacion ---

def verificarCiudades(tener, tiene):	# Verificar que la lista "tiene" tiene todas las paradas en "tener" en el mismo orden
	for i in range(len(tener)):
		if (tener[0] != tiene[0]): return i
		
		del tener[0]
		del tiene[0]
		
		if (len(tiene) == 0 or len(tener) == 0): break
	
	return None

# --- Creacion de linea ---

def crearLinea(i, ciuds):
	print("Creando hoja para " + lineas[i].nombre)
	nuevoTren = workbook.create_sheet(title = lineas[i].nombre)
	
	# Añadir el primer horario
	nuevoTren.cell(row = 1, column = 1, value = 2)
	nuevoTren.cell(row = 1, column = 2, value = 3)
	nuevoTren.cell(row = 1, column = 3, value = 4)
	nuevoTren.cell(row = 1, column = 4, value = 3 + len(ciuds))
	
	# Añadir el texto del tren
	negrita = Font(bold = True, size = 10, name = "Arial")
	
	nuevoTren.merge_cells("A2:C2")
	nuevoTren.cell(row = 2, column = 1, value = lineas[i].nombre).font = negrita
	
	nuevoTren.cell(row = 3, column = 1, value = "Parada").font = negrita
	nuevoTren.cell(row = 3, column = 2, value = "Llegada").font = negrita
	nuevoTren.cell(row = 3, column = 3, value = "Salida").font = negrita
	
	# Añadir los datos del tren actual
	for c in range(len(ciuds)):
		nuevoTren.cell(row = c + 4, column = 1, value = ciuds[c]).font = negrita		# Añadir las paradas del tren
		
		nuevoTren.cell(row = c + 4, column = 2, value = lineas[i].trayectos[0][2][c])	# Añadir el horario actual de llegada
		nuevoTren.cell(row = c + 4, column = 3, value = lineas[i].trayectos[0][3][c])	# Añadir el horario actual de salida
	
	# Añadir el texto de la vuelta
	nuevoTren.cell(row = 1, column = 5, value = 2)
	nuevoTren.cell(row = 1, column = 6, value = 2)
	nuevoTren.cell(row = 1, column = 7, value = 9 + len(ciuds))
	nuevoTren.cell(row = 1, column = 8, value = 8 + (2 * len(ciuds)))
	
	nuevoTren.merge_cells('A' + str(7 + len(ciuds)) + ":C" + str(7 + len(ciuds)))
	nuevoTren.cell(row = 7 + len(ciuds), column = 1, value = lineas[i].nombre).font = negrita
	
	nuevoTren.cell(row = 8 + len(ciuds), column = 1, value = "Parada").font = negrita
	
	# Añadir las paradas para el horario de vuelta
	ciuds.reverse()
	for c in range(len(ciuds)):	nuevoTren.cell(row = c + 9 + len(ciuds), column = 1, value = ciuds[c]).font = negrita # Añadir las paradas del tren
	
	workbook.save(horariosF)	# Guardar el archivo

def procesarTrenes():
	for i in tqdm(range(len(lineas)), unit = "Lineas"):		# Bucle que pasa por todos los trenes ya existentes
		
		# . Crear una nueva linea (temp) vacia .
		
		lineas_escribir.append(Linea(lineas[i].id, lineas[i].nombre, lineas[i].color, [Trayecto([], [], [])], [Ruta(0, [])]))
		
		# . Coger los datos del editor como ejemplo .
		
		trayectoIda = copy.deepcopy(lineas[i].trayectos[0])	# Coger el primer horario que deberia ser de ida
		rutaIda = copy.deepcopy(lineas[i].rutas[0]) # Coger la primera ruta que deberia ser la de ida
		
		# . Meter los datos originales en la nueva linea .
		
		lineas_escribir[i].trayectos[0] = copy.deepcopy(trayectoIda)
		lineas_escribir[i].rutas[0] = copy.deepcopy(rutaIda)
		
		# Coger el nombre de las ciudades
		ciudades_trayecto = copy.deepcopy(trayectoIda.puntos)
		for c in range(len(ciudades_trayecto)):
			ciudades_trayecto[c] = getCiudadfromId(ciudades_trayecto[c].id_ciudad).nombre.upper()
		
		# Coger la línea del tren desde el Excel si existe, si no, crearla
		try:
			sheet = workbook[lineas[i].nombre]
		except KeyError:
			crearLinea(i, copy.deepcopy(ciudades))
			sheet = workbook.create_sheet(lineas[i].nombre)	# Crear la hoja si no existe
			continue
		
		# --- Crear la base para el viaje de vuelta ---
		
		# . Trayecto .
		
		trayectoVuelta = copy.deepcopy(trayectoIda)	# Los trayectos de vuelta son iguales que los de ida, excepto que se invierte el orden de las paradas
		trayectoVuelta.puntos = trayectoVuelta.puntos[::-1] # Invertir las paradas
		
		# . Ruta .
		
		rutaVuelta = copy.deepcopy(rutaIda)
		rutaVuelta.puntos = rutaVuelta.puntos[::-1]
		
		# . Placeholder Trayecto Class .
		
		curr = copy.deepcopy(trayectoIda)
		
		# --- Leer valores para empezar el trayecto de ida ---
		
		start_col = sheet.cell(row = 1, column = 1).value
		end_col = sheet.cell(row = 1, column = 2).value
		start_row = sheet.cell(row = 1, column = 3).value
		end_row = sheet.cell(row = 1, column = 4).value
		
		# . Verificar ciudades de ida .
		
		res = verificarCiudades(copy.deepcopy(ciudades_trayecto), [str(sheet.cell(row = r, column = 1).value).upper() for r in range(start_row, end_row + 1)])
		
		if (res != None):
			print("Ida: El nombre de las ciudades para la línea", sheet.title, "no concuerda. Se esperaba", ciudades_trayecto[res], "se encontró", sheet.cell(row = start_row + res, column = 1).value)
			continue
		
		if (end_row + 1 - start_row != len(curr.puntos)):
			print("Ida: ", sheet.title, end_row + 1 - start_row , "paradas en el Excel, se esperaban", len(curr.puntos), "paradas")
			continue
		
		for col in range(start_col, end_col, 2):
			curr = copy.deepcopy(trayectoIda)	# Resetear el horario de ida por si se han borrado paradas
			
			tiempoLleg = []
			tiempoSali = []
			
			for row in range(start_row, end_row + 1):
				llegada = int2hora(sheet.cell(row = row, column = col).value)
				salida = int2hora(sheet.cell(row = row, column = col + 1).value)
				
				if (salida <= llegada):
					print("Ida: Hora incorrecta para " + sheet.title + ", la hora de salida", salida, "es inferior o igual a la hora de llegada", llegada)
					break
				elif ( (llegada == None and salida != None) or (llegada != None and salida == None)):
					print("Ida: La salida o la llegada son None")
					break
				
				tiempoLleg.append(llegada)
				tiempoSali.append(salida)
				
			curr.llegadas = copy.deepcopy(tiempoLleg)
			curr.salidas = copy.deepcopy(tiempoSali)
			
			lineas_escribir[i].trayectos.append(copy.deepcopy(curr))
			lineas_escribir[i].rutas.append(copy.deepcopy(rutaIda))
		
		# Leer valores para empezar el trayecto de vuelta
		start_col = sheet.cell(row = 1, column = 5).value
		end_col = sheet.cell(row = 1, column = 6).value
		start_row = sheet.cell(row = 1, column = 7).value
		end_row = sheet.cell(row = 1, column = 8).value
		
		curr = copy.deepcopy(trayectoVuelta)
		
		# Verificar ciudades de ida
		ciudades_trayecto.reverse()
		res = verificarCiudades(copy.deepcopy(ciudades_trayecto), [str(sheet.cell(row = r, column = 1).value).upper() for r in range(start_row, end_row + 1)])
		
		if (res != None):
			print("Vuelta: El nombre de las ciudades para la línea", sheet.title, "no concuerda. Se esperaba", ciudades_trayecto[res], "se encontró", sheet.cell(row = start_row + res, column = 1).value)
			continue
		
		if (end_row + 1 - start_row != len(curr.puntos)):
			print("Vuelta: ", sheet.title, end_row + 1 - start_row , "paradas en el Excel, se esperaban", len(curr.puntos), "paradas")
			continue
		
		for col in range(start_col, end_col, 2):
			curr = copy.deepcopy(trayectoVuelta)	# Resetear el horario de vuelta por si se han borrado paradas
				
			tiempoLleg = []
			tiempoSali = []
			
			for row in range(start_row, end_row + 1):
				llegada = int2hora(sheet.cell(row = row, column = col).value)
				salida  = int2hora(sheet.cell(row = row, column = col + 1).value)
				
				
				if (salida <= llegada):
					print("Vuelta: Hora incorrecta para " + sheet.title + ", la hora de salida", salida, "es inferior o igual a la hora de llegada", llegada)
					break
				elif ( (llegada == None and salida != None) or (llegada != None and salida == None)):
					print("Vuelta: La salida O la llegada son None")
					break
				
				tiempoLleg.append(llegada)
				tiempoSali.append(salida)
			
			curr.llegadas = copy.deepcopy(tiempoLleg)
			curr.salidas = copy.deepcopy(tiempoSali)
			
			lineas_escribir[i].trayectos.append(copy.deepcopy(curr))
			lineas_escribir[i].rutas.append(copy.deepcopy(rutaVuelta))

# --------- INIT FUNCTS ---------

cargar_ciudades()
cargar_lineas()

workbook = openpyxl.load_workbook(horariosF, data_only = True) # Cargar los horarios en el Excel

if (len(lineas) > 0 and len(ciudades) > 0): # Sanity check
	procesarTrenes()
else:
	quit()

# --------- GUARDAR TRENES ---------

with open("LineasIC_output.json", "w+") as f:
	json.dump([l.__dict__ for l in lineas_escribir], f, indent = 4, default=lambda o: o.__dict__)
	f.close()

print("Trenes guardados")